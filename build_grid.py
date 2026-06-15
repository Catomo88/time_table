# -*- coding: utf-8 -*-
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.cell.cell import MergedCell

KFONT="맑은 고딕"
DAYS=["월","화","수","목","금","토","일"]
WHO=["첫째","둘째"]
A=True
S=[
 ["첫째","월","08:20","08:45","등교","돌봄/픽업","집→학교",False],
 ["첫째","월","08:45","13:30","정규수업","정규수업","○○초",False],
 ["첫째","월","13:50","15:20","방과후 회화미술","방과후","학교",False],
 ["첫째","월","15:50","16:20","한글수업","학원","한글교실",False],
 ["첫째","화","08:20","08:45","등교","돌봄/픽업","집→학교",False],
 ["첫째","화","08:45","13:30","정규수업","정규수업","○○초",False],
 ["첫째","화","14:00","15:00","태권도","학원","태권도장",False,"둘째와 함께 · 이모님(첫째 하교→둘째 하원→태권도)"],
 ["첫째","화","16:00","16:30","피아노","학원","피아노학원",False],
 ["첫째","수","08:20","08:45","등교","돌봄/픽업","집→학교",False],
 ["첫째","수","08:45","12:50","정규수업(단축)","정규수업","○○초",False],
 ["첫째","수","13:00","14:00","태권도","학원","태권도장",False],
 ["첫째","목","08:20","08:45","등교","돌봄/픽업","집→학교",False],
 ["첫째","목","08:45","13:30","정규수업","정규수업","○○초",False],
 ["첫째","목","13:40","15:10","방과후 메이키메이커","방과후","학교",False],
 ["첫째","목","17:10","17:40","한글수업","학원","한글교실",False],
 ["첫째","금","08:20","08:45","등교","돌봄/픽업","집→학교",False],
 ["첫째","금","08:45","12:50","정규수업(단축)","정규수업","○○초",False],
 ["첫째","금","12:50","14:20","방과후 종이접기","방과후","학교",False],
 ["첫째","금","15:00","16:00","미술학원","학원","미술학원",False,"새로 추가"],
 ["첫째","금","16:00","17:00","태권도","학원","태권도장",False,"둘째와 함께 · 이모님 픽업"],
 ["둘째","월","09:10","13:50","어린이집","정규수업","어린이집",False],
 ["둘째","월","16:20","16:50","한글수업","학원","한글교실",False],
 ["둘째","화","09:10","13:50","어린이집","정규수업","어린이집",False,"태권도 위해 일찍 하원 (확인 필요)"],
 ["둘째","화","14:00","15:00","태권도","학원","태권도장",False,"첫째와 함께 · 이모님 픽업"],
 ["둘째","수","09:10","13:50","어린이집","정규수업","어린이집",False],
 ["둘째","수","16:30","17:00","한글수업","학원","한글교실",False],
 ["둘째","목","09:10","13:50","어린이집","정규수업","어린이집",False],
 ["둘째","금","09:10","15:30","어린이집","정규수업","어린이집",False,"하원 시간 늘림"],
 ["둘째","금","16:00","17:00","태권도","학원","태권도장",False,"첫째와 함께 · 이모님 픽업"],
 ["첫째","수","19:40","20:10","첼로","학원","첼로교실",False,"수/토 중 하나 (주말 일정에 따라 변동)"],
 ["첫째","토","09:30","10:00","첼로","학원","첼로교실",False,"수/토 중 하나 (주말 일정에 따라 변동)"],
 ["첫째","일","15:00","16:00","수영","학원","수영장",False,""],
 ["둘째","목","14:30","15:00","축구","학원","축구교실",False,""],
]
S=[r+[''] if len(r)==8 else r for r in S]
CATCODE={"정규수업":"school","방과후":"after","학원":"academy","돌봄/픽업":"care","기타":"etc"}
FILL={"정규수업":"CFE2F3","방과후":"D9EAD3","학원":"FCE5CD","돌봄/픽업":"FFF2CC","기타":"E6E0EC"}
TXT ={"정규수업":"1F4E79","방과후":"2E6B1F","학원":"9C4A16","돌봄/픽업":"7A5E00","기타":"5B3F7A"}
HEAD={"첫째":"2E8B57","둘째":"2F6FC0"}
toMin=lambda t:int(t[:2])*60+int(t[3:])
GS,STEP=480,15
_maxend=max((toMin(r[3]) for r in S), default=1080)
GE=max(1080, ((_maxend//30)+1)*30)   # 막내 일정 끝까지 + 여유, 최소 18:00
NROWS=(GE-GS)//STEP              # 20
thin=Side(style="thin",color="C9C9C9"); med=Side(style="medium",color="9AA0A6")
border=Border(left=thin,right=thin,top=thin,bottom=thin)

wb=Workbook(); wb.remove(wb.active)

def grid_sheet(who):
    ws=wb.create_sheet(f"{who} 주간표")
    ws.sheet_view.showGridLines=False
    t=ws.cell(1,1,f"{who} 주간 시간표  (2026.03)")
    t.font=Font(name=KFONT,bold=True,size=16,color="FFFFFF")
    t.fill=PatternFill("solid",fgColor=HEAD[who])
    t.alignment=Alignment("center","center")
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=1+len(DAYS))
    ws.row_dimensions[1].height=30
    # header row 2
    hc=ws.cell(2,1,"시간"); hc.font=Font(name=KFONT,bold=True,color="FFFFFF")
    hc.fill=PatternFill("solid",fgColor="595959"); hc.alignment=Alignment("center","center"); hc.border=border
    for j,d in enumerate(DAYS,2):
        c=ws.cell(2,j,d); c.font=Font(name=KFONT,bold=True,color="FFFFFF",size=12)
        c.fill=PatternFill("solid",fgColor="404040"); c.alignment=Alignment("center","center"); c.border=border
    ws.row_dimensions[2].height=20
    # time rows
    for i in range(NROWS):
        r=3+i; m=GS+i*STEP
        lab=f"{m//60:02d}:{m%60:02d}" if m%30==0 else ""
        c=ws.cell(r,1,lab); c.font=Font(name=KFONT,size=8,bold=(m%60==0),color="555555")
        c.alignment=Alignment("center","center"); c.border=border
        c.fill=PatternFill("solid",fgColor="EFEFEF" if m%60==0 else "F7F7F7")
        ws.row_dimensions[r].height=15
        for j in range(2,2+len(DAYS)):
            cc=ws.cell(r,j); cc.border=border
    # place activities
    for row in S:
        if row[0]!=who: continue
        _,day,s,e,title,cat,place,assumed,note=row
        if day not in DAYS: continue
        col=2+DAYS.index(day)
        sm,em=toMin(s),toMin(e)
        sr=3+max(0,(sm-GS)//STEP)
        er=3+min(NROWS,-(-(em-GS)//STEP))-1   # ceil end, inclusive last row
        if er<sr: er=sr
        while sr<=er and isinstance(ws.cell(sr,col),MergedCell): sr+=1
        for r in range(sr,er+1):
            cell=ws.cell(r,col)
            cell.fill=PatternFill("solid",fgColor=FILL[cat])
            cell.border=Border(left=thin,right=thin,top=thin,bottom=thin)
        top=ws.cell(sr,col)
        txt=f"{title}\n{s}–{e}"+("*" if assumed else "")
        if note: txt+=f"\n{note.split(' · ')[0]}"
        elif place and cat in ("학원","방과후"): txt+=f"\n{place}"
        top.value=txt
        top.font=Font(name=KFONT,size=10,bold=True,color=TXT[cat])
        top.alignment=Alignment("center","center",wrap_text=True)
        if er>sr: ws.merge_cells(start_row=sr,start_column=col,end_row=er,end_column=col)
        # thicker outline around block
        for r in range(sr,er+1):
            cur=ws.cell(r,col); l=border.left;rt=border.right
            cur.border=Border(left=med,right=med,
                              top=med if r==sr else thin, bottom=med if r==er else thin)
    # widths / print
    ws.column_dimensions["A"].width=7
    for j in range(2,2+len(DAYS)):
        ws.column_dimensions[chr(64+j)].width=17
    ws.print_area=f"A1:{chr(64+1+len(DAYS))}{2+NROWS}"
    ws.page_setup.orientation="landscape"
    ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=1
    ws.sheet_properties.pageSetUpPr.fitToPage=True
    ws.page_margins.left=ws.page_margins.right=0.3
    ws.page_margins.top=ws.page_margins.bottom=0.4
    # legend under grid
    lr=3+NROWS+1
    ws.cell(lr,1,"색상:").font=Font(name=KFONT,bold=True,size=9)
    for k,(catn,fc) in enumerate(FILL.items()):
        cc=ws.cell(lr,2+k,catn); cc.fill=PatternFill("solid",fgColor=fc)
        cc.font=Font(name=KFONT,size=9,color=TXT[catn]); cc.alignment=Alignment("center","center"); cc.border=border
    ws.cell(lr+1,1,"* 표시 = 사진에 끝나는 시간이 없어 약 50분으로 추정한 항목 (확인 필요)").font=Font(name=KFONT,italic=True,size=8,color="C00000")

for w in WHO: grid_sheet(w)

# ---- 일정 (editable list) ----
ws=wb.create_sheet("일정(편집용)"); ws.sheet_view.showGridLines=False
t=ws.cell(1,1,"일정 편집용 — 여기 고친 뒤 '주간표 갱신' 요청하면 위 시간표가 다시 그려집니다")
t.font=Font(name=KFONT,bold=True,size=11,color="FFFFFF"); t.fill=PatternFill("solid",fgColor="595959")
ws.merge_cells("A1:H1"); ws.row_dimensions[1].height=24
heads=["구분","요일","시작","종료","활동","분류","장소","비고"]
for j,h in enumerate(heads,1):
    c=ws.cell(2,j,h); c.font=Font(name=KFONT,bold=True,color="FFFFFF")
    c.fill=PatternFill("solid",fgColor="808080"); c.alignment=Alignment("center","center"); c.border=border
for i,row in enumerate(S):
    r=3+i
    bigo=row[8] if (len(row)>8 and row[8]) else ""
    if row[7]: bigo=(bigo+" / " if bigo else "")+"※ 종료시간 추정"
    vals=[row[0],row[1],row[2],row[3],row[4],row[5],row[6],bigo]
    for j,v in enumerate(vals,1):
        c=ws.cell(r,j,v); c.font=Font(name=KFONT,size=10); c.border=border
        c.alignment=Alignment("center","center") if j in(1,2,3,4,6) else Alignment("left","center")
    ws.cell(r,6).fill=PatternFill("solid",fgColor=FILL[row[5]])
    ws.cell(r,1).font=Font(name=KFONT,size=10,bold=True,color=HEAD[row[0]])
for col,w in {"A":7,"B":6,"C":7,"D":7,"E":24,"F":12,"G":14,"H":16}.items():
    ws.column_dimensions[col].width=w
ws.freeze_panes="A3"
dvw=DataValidation(type="list",formula1='"'+",".join(WHO)+'"',allow_blank=True)
dvd=DataValidation(type="list",formula1='"'+",".join(["월","화","수","목","금","토","일"])+'"',allow_blank=True)
dvc=DataValidation(type="list",formula1='"'+",".join(["정규수업","방과후","학원","돌봄/픽업","기타"])+'"',allow_blank=True)
for dv,c in [(dvw,"A"),(dvd,"B"),(dvc,"F")]:
    ws.add_data_validation(dv); dv.add(f"{c}3:{c}200")

wb.save("/sessions/sweet-beautiful-hawking/mnt/outputs/시간표.xlsx")
data=[{"who":r[0],"day":r[1],"s":r[2],"e":r[3],"title":r[4],"cat":CATCODE[r[5]],"place":r[6],"assumed":r[7],"note":r[8]} for r in S]
open("/sessions/sweet-beautiful-hawking/mnt/outputs/_data.json","w").write(json.dumps(data,ensure_ascii=False))
print("grid xlsx saved, rows",len(data))
