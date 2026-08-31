# -*- coding: utf-8 -*-
"""첫째(서원) 주간 시간표를 A4 한 장 카드로 만든다. -> 서원_주간카드.xlsx"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

KF="맑은 고딕"; WHO="첫째"; NAME="서원"
DAYS=["월","화","수","목","금","토","일"]
FILL={"정규수업":"DDE5F9","방과후":"C3D0F1","학원":"A9BAEB","돌봄/픽업":"EFF3FD","기타":"E4E9F8"}
TXT ={"정규수업":"20388C","방과후":"1B2E77","학원":"14225C","돌봄/픽업":"46589E","기타":"20388C"}
thin=Side(style="thin",color="D9DCE3"); border=Border(left=thin,right=thin,top=thin,bottom=thin)

S=[x for x in json.load(open("schedule.json",encoding="utf-8")) if x["who"]==WHO]
by={d:sorted([x for x in S if x["day"]==d],key=lambda r:r["s"]) for d in DAYS}
after={d:[x for x in by[d] if x["cat"] not in ("정규수업","돌봄/픽업")] for d in DAYS}
NCHIP=max(3,max(len(v) for v in after.values()))

wb=Workbook(); ws=wb.active; ws.title=f"{NAME} 주간카드"
ws.sheet_view.showGridLines=False
NC=2+NCHIP

t=ws.cell(1,1,f"{NAME} 주간 시간표")
t.font=Font(name=KF,bold=True,size=26,color="FFFFFF")
t.fill=PatternFill("solid",fgColor="3B4CB8"); t.alignment=Alignment("center","center")
ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=NC); ws.row_dimensions[1].height=46

sub=ws.cell(2,1,"매주 반복되는 일정 · 등교 08:20  |  학교 안 수업(방과후·맞춤형)과 학원만 정리했어요")
sub.font=Font(name=KF,size=11,color="6B7078"); sub.alignment=Alignment("center","center")
ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=NC); ws.row_dimensions[2].height=26

hs=[("요일",9),("학교",26)]+[(f"일정 {i+1}",28) for i in range(NCHIP)]
for j,(h,wd) in enumerate(hs,1):
    c=ws.cell(3,j,h); c.font=Font(name=KF,bold=True,size=11,color="FFFFFF")
    c.fill=PatternFill("solid",fgColor="6B7280"); c.alignment=Alignment("center","center"); c.border=border
    ws.column_dimensions[get_column_letter(j)].width=wd
ws.row_dimensions[3].height=24

for i,d in enumerate(DAYS):
    r=4+i; ws.row_dimensions[r].height=62
    wk=i>4
    dc=ws.cell(r,1,d)
    dc.font=Font(name=KF,bold=True,size=17,color="C0392B" if wk else "23262D")
    dc.fill=PatternFill("solid",fgColor="FBF7F5" if wk else "F4F5F8")
    dc.alignment=Alignment("center","center"); dc.border=border

    sc=[x for x in by[d] if x["cat"]=="정규수업"]
    cell=ws.cell(r,2)
    if sc:
        x=sc[0]; cell.value=f"{x['s']}–{x['e']}\n{x['title'].replace('정규수업','학교')}"
        cell.fill=PatternFill("solid",fgColor=FILL["정규수업"]); cell.font=Font(name=KF,size=12,bold=True,color=TXT["정규수업"])
    else:
        cell.value="—"; cell.font=Font(name=KF,size=12,color="B6BAC2")
    cell.alignment=Alignment("center","center",wrap_text=True); cell.border=border

    for k in range(NCHIP):
        c=ws.cell(r,3+k); c.border=border; c.alignment=Alignment("center","center",wrap_text=True)
        if k<len(after[d]):
            x=after[d][k]
            nm=x["title"].replace("방과후 ","").replace("맞춤형 ","")
            tag="방과후 " if x["title"].startswith("방과후") else ("맞춤형 " if x["title"].startswith("맞춤형") else "")
            c.value=f"{x['s']}–{x['e']}\n{tag}{nm}"+(f"\n{x['place']}" if x.get("place") else "")
            c.fill=PatternFill("solid",fgColor=FILL[x["cat"]])
            c.font=Font(name=KF,size=11,bold=True,color=TXT[x["cat"]])
        elif not after[d]:
            if k==0:
                c.value="일정 없음"; c.font=Font(name=KF,size=11,color="B6BAC2")

lr=4+len(DAYS)
lg=ws.cell(lr,1,"색상"); lg.font=Font(name=KF,bold=True,size=9,color="6B7078")
for k,(cat,nm) in enumerate([("정규수업","학교"),("방과후","방과후·맞춤형"),("학원","학원")]):
    c=ws.cell(lr,2+k,nm); c.fill=PatternFill("solid",fgColor=FILL[cat])
    c.font=Font(name=KF,size=9,bold=True,color=TXT[cat]); c.alignment=Alignment("center","center"); c.border=border
ws.row_dimensions[lr].height=18

ws.print_area=f"A1:{get_column_letter(NC)}{lr}"
ws.page_setup.orientation="landscape"; ws.page_setup.paperSize=9
ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=1
ws.sheet_properties.pageSetUpPr.fitToPage=True
ws.page_margins.left=ws.page_margins.right=0.3
ws.page_margins.top=ws.page_margins.bottom=0.4
wb.save("서원_주간카드.xlsx")
print("saved 서원_주간카드.xlsx / 일정칸",NCHIP)
